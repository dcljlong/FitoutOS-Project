"""
Generate test programme files for edge case testing.
Covers:
- missing columns
- mixed date formats  
- blank rows
- duplicate rows
- merged-cell style input (simulated)
- missing predecessor
- missing duration
"""
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import csv
from datetime import datetime, timedelta

TEST_DIR = os.path.join(os.path.dirname(__file__), "test_programmes")
os.makedirs(TEST_DIR, exist_ok=True)

def create_standard_programme():
    """Standard valid programme for baseline comparison"""
    data = {
        "Task ID": ["T001", "T002", "T003", "T004", "T005"],
        "Task Name": ["Wall Framing", "Wall Linings", "Stopping", "Ceiling Framing", "Ceiling Linings"],
        "Start Date": ["2026-01-15", "2026-01-20", "2026-01-27", "2026-01-22", "2026-02-01"],
        "Finish Date": ["2026-01-19", "2026-01-26", "2026-02-02", "2026-01-28", "2026-02-05"],
        "Duration": [5, 5, 5, 5, 5],
        "Predecessor": ["", "T001", "T002", "T001", "T004"],
        "Trade": ["Framing", "Linings", "Stopping", "Ceilings", "Ceilings"]
    }
    df = pd.DataFrame(data)
    path = os.path.join(TEST_DIR, "01_standard_programme.xlsx")
    df.to_excel(path, index=False)
    print(f"Created: {path}")
    return path

def create_missing_columns_programme():
    """Programme with missing required columns (no Duration column)"""
    data = {
        "Task ID": ["T001", "T002", "T003"],
        "Task Name": ["Wall Framing", "Wall Linings", "Stopping"],
        "Start Date": ["2026-01-15", "2026-01-20", "2026-01-27"],
        "Finish Date": ["2026-01-19", "2026-01-26", "2026-02-02"],
        # Duration column intentionally missing
        "Predecessor": ["", "T001", "T002"],
    }
    df = pd.DataFrame(data)
    path = os.path.join(TEST_DIR, "02_missing_columns.xlsx")
    df.to_excel(path, index=False)
    print(f"Created: {path}")
    return path

def create_mixed_date_formats():
    """Programme with mixed date formats"""
    data = {
        "Task ID": ["T001", "T002", "T003", "T004", "T005", "T006"],
        "Task Name": ["Wall Framing", "Wall Linings", "Stopping", "Ceiling Framing", "Insulation", "Final Fix"],
        "Start Date": [
            "2026-01-15",           # ISO format
            "15/01/2026",           # UK format
            "01/15/2026",           # US format
            "Jan 22, 2026",         # Long format
            "2026/01/29",           # Alternative ISO
            datetime(2026, 2, 5),   # Python datetime object
        ],
        "Finish Date": [
            "2026-01-19",
            "19/01/2026", 
            "01/19/2026",
            "Jan 28, 2026",
            "2026/02/04",
            datetime(2026, 2, 10),
        ],
        "Duration": [5, 5, None, 5, 5, 5],  # One None to test calculation
        "Predecessor": ["", "T001", "T002", "T001", "T004", "T005"],
    }
    df = pd.DataFrame(data)
    path = os.path.join(TEST_DIR, "03_mixed_date_formats.xlsx")
    df.to_excel(path, index=False)
    print(f"Created: {path}")
    return path

def create_blank_rows_programme():
    """Programme with blank/empty rows interspersed"""
    data = {
        "Task ID": ["T001", None, "T002", "", "T003", None, "T004"],
        "Task Name": ["Wall Framing", None, "Wall Linings", "", "Stopping", None, "Ceiling Framing"],
        "Start Date": ["2026-01-15", None, "2026-01-20", "", "2026-01-27", None, "2026-01-22"],
        "Finish Date": ["2026-01-19", None, "2026-01-26", "", "2026-02-02", None, "2026-01-28"],
        "Duration": [5, None, 5, "", 5, None, 5],
        "Predecessor": ["", None, "T001", "", "T002", None, "T001"],
    }
    df = pd.DataFrame(data)
    path = os.path.join(TEST_DIR, "04_blank_rows.xlsx")
    df.to_excel(path, index=False)
    print(f"Created: {path}")
    return path

def create_duplicate_rows_programme():
    """Programme with duplicate task rows"""
    data = {
        "Task ID": ["T001", "T002", "T002", "T003", "T001", "T003", "T004"],  # Duplicates
        "Task Name": ["Wall Framing", "Wall Linings", "Wall Linings", "Stopping", "Wall Framing", "Stopping", "Ceilings"],
        "Start Date": ["2026-01-15", "2026-01-20", "2026-01-20", "2026-01-27", "2026-01-15", "2026-01-27", "2026-02-03"],
        "Finish Date": ["2026-01-19", "2026-01-26", "2026-01-26", "2026-02-02", "2026-01-19", "2026-02-02", "2026-02-07"],
        "Duration": [5, 5, 5, 5, 5, 5, 5],
        "Predecessor": ["", "T001", "T001", "T002", "", "T002", "T003"],
    }
    df = pd.DataFrame(data)
    path = os.path.join(TEST_DIR, "05_duplicate_rows.xlsx")
    df.to_excel(path, index=False)
    print(f"Created: {path}")
    return path

def create_merged_cell_style():
    """Programme with merged-cell style (phase headers spanning multiple rows)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Programme"
    
    # Header row
    headers = ["Phase", "Task ID", "Task Name", "Start", "Finish", "Duration", "Predecessor"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    
    # Data with phase groupings (simulating merged cells by repeating phase)
    data = [
        ["Preliminaries", "T001", "Site Measure", "2026-01-10", "2026-01-10", 1, ""],
        ["Preliminaries", "T002", "Material Delivery", "2026-01-11", "2026-01-12", 2, "T001"],
        ["", "", "", "", "", "", ""],  # Blank row
        ["Construction", "T003", "Wall Framing", "2026-01-13", "2026-01-17", 5, "T002"],
        ["Construction", "T004", "Wall Linings", "2026-01-18", "2026-01-24", 5, "T003"],
        ["Construction", "T005", "Stopping", "2026-01-25", "2026-01-31", 5, "T004"],
        ["", "", "", "", "", "", ""],  # Blank row
        ["Finishing", "T006", "Paint Prep", "2026-02-01", "2026-02-03", 3, "T005"],
    ]
    
    for row_idx, row in enumerate(data, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    
    path = os.path.join(TEST_DIR, "06_merged_cell_style.xlsx")
    wb.save(path)
    print(f"Created: {path}")
    return path

def create_missing_predecessor_programme():
    """Programme with missing/invalid predecessors"""
    data = {
        "Task ID": ["T001", "T002", "T003", "T004", "T005"],
        "Task Name": ["Wall Framing", "Wall Linings", "Stopping", "Ceiling Framing", "Final Fix"],
        "Start Date": ["2026-01-15", "2026-01-20", "2026-01-27", "2026-01-22", "2026-02-05"],
        "Finish Date": ["2026-01-19", "2026-01-26", "2026-02-02", "2026-01-28", "2026-02-10"],
        "Duration": [5, 5, 5, 5, 5],
        "Predecessor": [
            "",           # Valid - no predecessor
            "T001",       # Valid predecessor
            "T999",       # INVALID - doesn't exist
            None,         # Missing predecessor
            "T003,T004",  # Multiple predecessors (valid)
        ],
    }
    df = pd.DataFrame(data)
    path = os.path.join(TEST_DIR, "07_missing_predecessor.xlsx")
    df.to_excel(path, index=False)
    print(f"Created: {path}")
    return path

def create_missing_duration_programme():
    """Programme with missing/zero durations"""
    data = {
        "Task ID": ["T001", "T002", "T003", "T004", "T005", "T006"],
        "Task Name": ["Wall Framing", "Wall Linings", "Stopping", "Ceiling Framing", "Insulation", "Final Fix"],
        "Start Date": ["2026-01-15", "2026-01-20", "2026-01-27", "2026-01-22", "2026-01-29", "2026-02-05"],
        "Finish Date": ["2026-01-19", "2026-01-26", "2026-02-02", "2026-01-28", "2026-02-04", "2026-02-10"],
        "Duration": [5, None, 0, "", "5d", 5],  # Various invalid durations
        "Predecessor": ["", "T001", "T002", "T001", "T004", "T005"],
    }
    df = pd.DataFrame(data)
    path = os.path.join(TEST_DIR, "08_missing_duration.xlsx")
    df.to_excel(path, index=False)
    print(f"Created: {path}")
    return path

def create_messy_contractor_export():
    """Simulate messy real-world contractor export with multiple issues"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Construction Programme"
    
    # Add company header rows (like real contractor exports)
    ws.cell(row=1, column=1, value="ABC Construction Ltd")
    ws.cell(row=2, column=1, value="Project: Office Fitout")
    ws.cell(row=3, column=1, value="Generated: 25/03/2026")
    # Blank row
    ws.cell(row=5, column=1, value="")
    
    # Actual headers starting at row 6
    headers = ["ID", "Activity Description", "Start", "End", "Dur.", "Deps", "Notes", "Responsible"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=6, column=col, value=h)
    
    # Messy data with various issues
    data = [
        ["1", "PRELIMINARIES", "", "", "", "", "SECTION HEADER", ""],  # Section header row
        ["1.1", "Site establishment", "13-Jan-26", "14-Jan-26", "2d", "", "", "MC"],
        ["1.2", "Material delivery", "15/01/2026", "15/01/2026", 1, "1.1", "", "Subbie"],
        ["", "", "", "", "", "", "", ""],  # Blank row
        ["2", "FRAMING WORKS", "", "", "", "", "SECTION HEADER", ""],  # Section header
        ["2.1", "Wall framing - Level 1", "2026-01-16", "2026-01-20", "5", "1.2", "", "Internal"],
        ["2.2", "Wall framing - Level 2", "Jan 21, 2026", "Jan 25, 2026", 5, "2.1", "Depends on 2.1", "Internal"],
        ["2.3", "Ceiling framing", "26/01/26", "30/01/26", None, "2.2", "", "Internal"],  # Missing duration
        ["", "", "", "", "", "", "", ""],  # Another blank row
        ["3", "LININGS", "", "", "", "", "SECTION HEADER", ""],
        ["3.1", "Gib fixing - walls", "31-Jan-2026", "06-Feb-2026", "5 days", "2.1,2.2", "Multiple deps", "Internal"],
        ["3.2", "Gib fixing - ceilings", "07/02/2026", "12/02/2026", "4", "3.1,2.3", "", "Internal"],
        ["3.1", "Gib fixing - walls", "31-Jan-2026", "06-Feb-2026", "5 days", "2.1,2.2", "DUPLICATE", "Internal"],  # Duplicate
        ["", "", "", "", "", "", "", ""],
        ["4", "STOPPING & FINISHING", "", "", "", "", "", ""],
        ["4.1", "Stopping coat 1", "13/02/2026", "17/02/2026", 3, "3.1", "", "Subbies"],
        ["4.2", "Stopping coat 2", "18/02/2026", "20/02/2026", 2, "4.1", "", "Subbies"],
        ["4.3", "Final stopping", "21/02/2026", "25/02/2026", "", "UNKNOWN", "", "Subbies"],  # Invalid predecessor
    ]
    
    for row_idx, row in enumerate(data, 7):  # Start at row 7
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    
    path = os.path.join(TEST_DIR, "09_messy_contractor_export.xlsx")
    wb.save(path)
    print(f"Created: {path}")
    return path

def create_csv_programme():
    """CSV format programme for testing CSV parsing"""
    path = os.path.join(TEST_DIR, "10_csv_programme.csv")
    with open(path, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["Task ID", "Task Name", "Start Date", "Finish Date", "Duration", "Predecessor"])
        writer.writerow(["T001", "Wall Framing", "2026-01-15", "2026-01-19", "5", ""])
        writer.writerow(["T002", "Wall Linings", "2026-01-20", "2026-01-26", "5", "T001"])
        writer.writerow(["", "", "", "", "", ""])  # Blank row
        writer.writerow(["T003", "Stopping", "2026-01-27", "2026-02-02", "5", "T002"])
        writer.writerow(["T004", "Ceiling Framing", "2026-01-22", "2026-01-28", "5", "T001"])
    print(f"Created: {path}")
    return path

def create_all_test_files():
    """Generate all test files"""
    files = []
    files.append(create_standard_programme())
    files.append(create_missing_columns_programme())
    files.append(create_mixed_date_formats())
    files.append(create_blank_rows_programme())
    files.append(create_duplicate_rows_programme())
    files.append(create_merged_cell_style())
    files.append(create_missing_predecessor_programme())
    files.append(create_missing_duration_programme())
    files.append(create_messy_contractor_export())
    files.append(create_csv_programme())
    
    print(f"\nCreated {len(files)} test files in {TEST_DIR}")
    return files

if __name__ == "__main__":
    create_all_test_files()
