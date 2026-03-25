import os
import uuid
from pathlib import Path
from datetime import datetime, timezone

from dotenv import load_dotenv
from pymongo import MongoClient
from openpyxl import load_workbook

ROOT = Path(r"D:\FitoutOS Project\app\backend")
load_dotenv(ROOT / ".env")

excel_path = Path(r"/mnt/data/Job Task Codes.xlsx")
if not excel_path.exists():
    excel_path = Path(r"D:\FitoutOS Project\app\backend\Job Task Codes.xlsx")

if not excel_path.exists():
    raise FileNotFoundError(f"Excel file not found at {excel_path}")

client = MongoClient(os.environ["MONGO_URL"])
db = client[os.environ["DB_NAME"]]

wb = load_workbook(excel_path, data_only=True)
ws = wb[wb.sheetnames[0]]  # Sheet1

rows = list(ws.iter_rows(values_only=True))
headers = [str(h).strip() if h is not None else "" for h in rows[0]]

code_idx = None
name_idx = None

for i, h in enumerate(headers):
    hl = h.lower()
    if code_idx is None and "code" in hl:
        code_idx = i
    if name_idx is None and ("name" in hl or "description" in hl or "label" in hl):
        name_idx = i

if code_idx is None:
    code_idx = 0
if name_idx is None:
    name_idx = 1

codes = []
for row in rows[1:]:
    if row is None:
        continue
    code = row[code_idx] if code_idx < len(row) else None
    name = row[name_idx] if name_idx < len(row) else None
    code = str(code).strip() if code is not None else ""
    name = str(name).strip() if name is not None else ""
    if not code or not name:
        continue
    codes.append({"code": code, "name": name})

dedup = {}
for item in codes:
    dedup[item["code"]] = item

codes = list(dedup.values())

backup_count = db.master_task_codes.count_documents({})
print({"existing_master_count": backup_count, "import_count": len(codes)})

db.master_task_codes.delete_many({})

now = datetime.now(timezone.utc).isoformat()
docs = []
for item in codes:
    docs.append({
        "id": str(uuid.uuid4()),
        "code": item["code"],
        "name": item["name"],
        "description": None,
        "category": None,
        "is_global_fallback": False,
        "is_active": True,
        "created_at": now,
    })

if docs:
    db.master_task_codes.insert_many(docs)

print({"inserted_master_count": len(docs)})
for row in db.master_task_codes.find({}, {"_id": 0, "code": 1, "name": 1}).sort("code", 1):
    print(f'{row["code"]} - {row["name"]}')
