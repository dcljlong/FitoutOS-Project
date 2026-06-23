"""
Robust Programme Parser Module for FitoutOS

Handles structured parsing of Excel, CSV, and XLS programme files.
Designed to handle common edge cases:
- Missing columns
- Mixed date formats
- Blank/empty rows
- Duplicate rows
- Merged cells (simulated via phase groupings)
- Missing predecessors
- Missing durations
- Messy contractor exports with headers/footers
"""

import os
import re
import csv
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path

import xlrd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Common column name mappings (normalized lowercase)
COLUMN_MAPPINGS = {
    # Task ID variations
    "task_id": ["task id", "taskid", "id", "task_id", "code", "item", "no", "no.", "wbs", "activity id", "act id"],
    # Task name variations
    "task_name": ["task name", "taskname", "name", "activity", "activity description", "description", "activity name", "task", "item description"],
    # Start date variations
    "start_date": ["start", "start date", "startdate", "planned start", "start_date", "begin", "begin date"],
    # End date variations  
    "end_date": ["finish", "finish date", "finishdate", "end", "end date", "enddate", "planned finish", "finish_date", "end_date", "complete"],
    # Duration variations
    "duration": ["duration", "dur", "dur.", "days", "length", "work days"],
    # Predecessor variations
    "predecessor": ["predecessor", "predecessors", "pred", "deps", "depends", "depends on", "dependency", "dependencies", "link", "prev"],
    # Phase variations
    "phase": ["phase", "section", "category", "stage", "group"],
    # Trade variations
    "trade": ["trade", "contractor", "responsible", "resource", "assigned", "owner"],
    # Notes variations
    "notes": ["notes", "note", "comments", "comment", "remarks"],
    # Project Plan / commercial programme review extras
    "hours_quoted": ["hrs quoted", "hours quoted", "quoted hours", "scope hours", "hrs"],
    "crew_label": ["crew label", "crew", "crew size", "crew_label"],
}

# Date format patterns to try (in order)
DATE_FORMATS = [
    "%Y-%m-%d",          # 2026-01-15
    # FITOUTOS / PROJECT PLAN 365 DATE FORMATS V3
    "%a %d/%m/%Y",        # Thu 15/01/2026
    "%a %d/%m/%y",         # Thu 15/01/26
    "%d/%m/%y",           # 15/01/26
    "%d/%m/%Y",          # 15/01/2026
    "%m/%d/%Y",          # 01/15/2026
    "%d-%m-%Y",          # 15-01-2026
    "%d-%b-%Y",          # 15-Jan-2026
    "%d-%b-%y",          # 15-Jan-26
    "%b %d, %Y",         # Jan 15, 2026
    "%Y/%m/%d",          # 2026/01/15
    "%d.%m.%Y",          # 15.01.2026
    "%d %b %Y",          # 15 Jan 2026
    "%d %B %Y",          # 15 January 2026
]


class ParseResult:
    """Result of parsing a programme file"""
    def __init__(self):
        self.items: List[Dict[str, Any]] = []
        self.warnings: List[Dict[str, str]] = []
        self.errors: List[Dict[str, str]] = []
        self.metadata: Dict[str, Any] = {
            "total_rows_read": 0,
            "blank_rows_skipped": 0,
            "duplicate_rows_found": 0,
            "header_row_found": None,
            "columns_detected": [],
            "missing_columns": [],
            "date_formats_found": set(),
        }
    
    def add_warning(self, row: int, message: str):
        self.warnings.append({"row": row, "message": message})
    
    def add_error(self, row: int, message: str):
        self.errors.append({"row": row, "message": message})
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            "items": self.items,
            "warnings": self.warnings,
            "errors": self.errors,
            "metadata": {
                **self.metadata,
                "date_formats_found": list(self.metadata["date_formats_found"])
            }
        }


def normalize_column_name(col_name: str) -> str:
    """Normalize a column name for matching"""
    if col_name is None:
        return ""
    return re.sub(r'[^a-z0-9\s]', '', str(col_name).lower().strip())


def detect_column_mapping(headers: List[str]) -> Dict[str, int]:
    """
    Detect which column index maps to which field.
    Returns a dict of field_name -> column_index.
    """
    mapping = {}
    normalized_headers = [normalize_column_name(h) for h in headers]
    
    for field, variations in COLUMN_MAPPINGS.items():
        for idx, header in enumerate(normalized_headers):
            if header in variations:
                mapping[field] = idx
                break
    
    return mapping


def parse_date(value: Any) -> Optional[str]:
    """
    Parse a date value from various formats.
    Returns ISO format string (YYYY-MM-DD) or None if parsing fails.
    """
    if value is None:
        return None
    
    # Handle datetime objects
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    
    # Handle Excel date serial numbers (float)
    if isinstance(value, (int, float)):
        try:
            # Excel serial date (days since 1899-12-30)
            if 30000 < value < 60000:  # Reasonable date range
                excel_epoch = datetime(1899, 12, 30)
                date = excel_epoch + timedelta(days=int(value))
                return date.strftime("%Y-%m-%d")
        except Exception:
            pass
    
    # Handle string values
    value_str = str(value).strip()
    if not value_str or value_str.lower() in ['', 'none', 'null', 'n/a', 'na', 'tbd']:
        return None
    
    # FITOUTOS / PROJECT PLAN 365 DATE NORMALISE V3
    # Project Plan 365 exports dates like "Thu 16/04/26".
    value_str = re.sub(r"\s+00:00:00$", "", value_str).strip()
    value_str = re.sub(r"\s+", " ", value_str).strip()
    value_str = re.sub(r"^(mon|tue|wed|thu|fri|sat|sun)\s+", "", value_str, flags=re.IGNORECASE).strip()
    # Try each date format
    for fmt in DATE_FORMATS:
        try:
            parsed = datetime.strptime(value_str, fmt)
            # Sanity check year
            if parsed.year < 2020 or parsed.year > 2050:
                continue
            return parsed.strftime("%Y-%m-%d")
        except ValueError:
            continue
    
    return None


def parse_duration(value: Any, start_date: Optional[str] = None, end_date: Optional[str] = None) -> Optional[float]:
    """
    Parse a duration value. Can also calculate from start/end dates.
    """
    # Try to extract numeric duration
    if value is not None:
        # Handle string with "d" or "days" suffix
        value_str = str(value).lower().strip()
        
        # Remove common suffixes
        for suffix in ['days', 'day', 'd', 'hrs', 'hours', 'hr', 'h']:
            value_str = value_str.replace(suffix, '').strip()
        
        try:
            duration = float(value_str)
            if duration > 0:
                return duration
        except ValueError:
            pass
    
    # Calculate from dates if duration is missing
    if start_date and end_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            end = datetime.strptime(end_date, "%Y-%m-%d")
            delta = (end - start).days + 1  # Include both start and end day
            if delta > 0:
                return float(delta)
        except Exception:
            pass
    
    return None


# FITOUTOS / PROJECT PLAN 365 ROW CLASSIFICATION V6
# Project Plan 365 exports often omit the ID column but keep predecessor
# references as source sequence numbers. Keep those rows reviewable, map
# plain/comma/FS lag references back to generated programme IDs, and preserve
# the source reference text for audit.
def split_predecessor_tokens(value: Any) -> List[str]:
    """Split predecessor text into Project Plan reference tokens without splitting lag text."""
    if value is None or str(value).strip() == "":
        return []

    value_str = str(value).strip()
    raw_tokens = re.split(r"[,;/]+", value_str)
    return [token.strip() for token in raw_tokens if token and token.strip()]


def normalise_project_predecessor_token(token: str) -> Optional[Dict[str, Any]]:
    """Return Project Plan predecessor parts, e.g. 27FS+10 days -> source_id 27."""
    if not token:
        return None

    token = str(token).strip()
    match = re.match(
        r"^\s*(?P<source_id>\d+)\s*(?P<link_type>FS|SS|FF|SF)?\s*(?P<lag>[+-]\s*\d+\s*(?:days?|day|d)?)?\s*$",
        token,
        flags=re.IGNORECASE,
    )
    if not match:
        return None

    link_type = (match.group("link_type") or "FS").upper()
    lag = match.group("lag")
    if lag:
        lag = re.sub(r"\s+", " ", lag.strip())

    return {
        "raw": token,
        "source_id": match.group("source_id"),
        "link_type": link_type,
        "lag": lag,
    }


def parse_predecessor(value: Any, valid_ids: set, source_id_map: Optional[Dict[str, str]] = None) -> Tuple[List[str], List[str]]:
    """
    Parse predecessor field. Supports direct IDs, comma lists, and Project Plan
    references such as 12FS-1 days or 27FS+10 days.
    Returns (valid_predecessors, invalid_predecessors).
    """
    tokens = split_predecessor_tokens(value)
    if not tokens:
        return [], []

    source_id_map = source_id_map or {}
    valid = []
    invalid = []

    for token in tokens:
        parsed = normalise_project_predecessor_token(token)
        if parsed:
            source_id = parsed["source_id"]
            mapped_id = source_id_map.get(source_id)
            if mapped_id:
                if mapped_id not in valid:
                    valid.append(mapped_id)
                continue

            if source_id in valid_ids:
                if source_id not in valid:
                    valid.append(source_id)
                continue

        if token in valid_ids:
            if token not in valid:
                valid.append(token)
        else:
            invalid.append(token)

    return valid, invalid


def _programme_is_empty_export_value(value: Any) -> bool:
    return value is None or str(value).strip().lower() in {"", "none", "null", "n/a", "na", "tbd", "0"}


def _programme_number_value(value: Any) -> Optional[float]:
    if value is None:
        return None

    value_str = str(value).strip().lower()
    if value_str in {"", "none", "null", "n/a", "na", "tbd"}:
        return None

    value_str = value_str.replace(",", "")
    value_str = re.sub(r"[^0-9.\-]", "", value_str)
    if value_str in {"", "-", "."}:
        return None

    try:
        return float(value_str)
    except ValueError:
        return None


def classify_programme_row(
    task_name: Any,
    predecessor_value: Any,
    contractor_value: Any,
    duration_value: Any,
    hours_quoted_value: Any,
    crew_label_value: Any,
    source_sequence: int,
) -> Tuple[str, bool]:
    """
    Classify a Project Plan export row without relying on Excel outline/indent
    metadata, because Project Plan 365 XLSX exports often flatten those values.
    """
    name = str(task_name or "").strip()
    name_lower = name.lower()
    pred_tokens = split_predecessor_tokens(predecessor_value)

    duration = parse_duration(duration_value)
    hours_quoted = _programme_number_value(hours_quoted_value)
    crew_label = _programme_number_value(crew_label_value)

    contractor_empty = _programme_is_empty_export_value(contractor_value)
    has_work_signal = (
        bool(pred_tokens)
        or not contractor_empty
        or (hours_quoted is not None and hours_quoted > 0)
        or (crew_label is not None and crew_label > 0)
    )

    if source_sequence == 1:
        return "project_summary", False

    if re.fullmatch(r"level\s+\d+[a-z]?", name_lower, flags=re.IGNORECASE):
        return "section", False

    if name.isupper() and any(ch.isalpha() for ch in name) and not pred_tokens:
        return "section", False

    if duration is not None and duration <= 0:
        return "milestone", False

    if not pred_tokens and contractor_empty:
        return "summary", False

    if not has_work_signal and duration is not None and duration > 1:
        return "summary", False

    return "task", True
def is_blank_row(row_values: List[Any]) -> bool:
    """Check if a row is effectively blank"""
    for val in row_values:
        if val is not None and str(val).strip():
            return False
    return True


def is_section_header_row(row_values: List[Any], col_mapping: Dict[str, int]) -> bool:
    """
    Detect if a row is a section header (e.g., "FRAMING WORKS", "LININGS")
    Section headers typically have:
    - Value in first column or task name column
    - No dates or durations
    """
    task_name_idx = col_mapping.get("task_name", col_mapping.get("task_id", 0))
    start_idx = col_mapping.get("start_date")
    end_idx = col_mapping.get("end_date")
    duration_idx = col_mapping.get("duration")
    
    # Get potential name value
    name_val = None
    if task_name_idx < len(row_values):
        name_val = row_values[task_name_idx]
    
    # Check if name exists but dates/duration are empty
    has_name = name_val and str(name_val).strip()
    has_start = start_idx is not None and start_idx < len(row_values) and row_values[start_idx] and str(row_values[start_idx]).strip()
    has_end = end_idx is not None and end_idx < len(row_values) and row_values[end_idx] and str(row_values[end_idx]).strip()
    has_duration = duration_idx is not None and duration_idx < len(row_values) and row_values[duration_idx] and str(row_values[duration_idx]).strip()
    
    # If has name but no scheduling info, likely a section header
    if has_name and not (has_start or has_end or has_duration):
        # Check if name looks like a header (all caps or short phrase)
        name_str = str(name_val).strip()
        if name_str.isupper() or len(name_str.split()) <= 3:
            return True
    
    return False


def find_header_row(rows: List[List[Any]]) -> Tuple[int, List[str]]:
    """
    Find the header row in a set of rows.
    Returns (row_index, header_list).
    """
    for idx, row in enumerate(rows[:20]):  # Check first 20 rows
        normalized = [normalize_column_name(c) for c in row]
        
        # Check how many known columns we find
        matches = 0
        for field, variations in COLUMN_MAPPINGS.items():
            for header in normalized:
                if header in variations:
                    matches += 1
                    break
        
        # If we find at least 3 matching columns, this is likely the header
        if matches >= 3:
            return idx, row
    
    # Default to first row if no clear header found
    return 0, rows[0] if rows else []


def read_xlsx_file(filepath: str) -> Tuple[List[List[Any]], str]:
    """Read an XLSX file and return rows + sheet name"""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    sheet_name = ws.title
    
    rows = []
    for row in ws.iter_rows(max_row=500, values_only=True):
        rows.append(list(row))
    
    wb.close()
    return rows, sheet_name


def read_xls_file(filepath: str) -> Tuple[List[List[Any]], str]:
    """Read an XLS file and return rows + sheet name"""
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)
    sheet_name = ws.name
    
    rows = []
    for row_idx in range(min(ws.nrows, 500)):
        row_values = ws.row_values(row_idx)
        rows.append(row_values)
    
    return rows, sheet_name


def read_csv_file(filepath: str) -> Tuple[List[List[Any]], str]:
    """Read a CSV file and return rows"""
    rows = []
    
    # Try to detect encoding and delimiter
    with open(filepath, 'r', encoding='utf-8-sig', errors='replace') as f:
        # Sniff delimiter
        sample = f.read(4096)
        f.seek(0)
        
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',;\t')
        except csv.Error:
            dialect = csv.excel
        
        reader = csv.reader(f, dialect)
        for row in reader:
            if len(rows) >= 500:
                break
            rows.append(row)
    
    return rows, "CSV"


def parse_programme_file(filepath: str) -> ParseResult:
    """
    Main entry point for parsing a programme file.
    Supports: .xlsx, .xls, .csv
    """
    result = ParseResult()
    filepath = str(filepath)
    ext = Path(filepath).suffix.lower()
    
    # Read file based on extension
    try:
        if ext == '.xlsx':
            rows, sheet_name = read_xlsx_file(filepath)
        elif ext == '.xls':
            rows, sheet_name = read_xls_file(filepath)
        elif ext == '.csv':
            rows, sheet_name = read_csv_file(filepath)
        else:
            result.add_error(0, f"Unsupported file format: {ext}")
            return result
    except Exception as e:
        result.add_error(0, f"Failed to read file: {str(e)}")
        return result
    
    if not rows:
        result.add_error(0, "File is empty or could not be read")
        return result
    
    # Find header row
    header_row_idx, headers = find_header_row(rows)
    result.metadata["header_row_found"] = header_row_idx + 1  # 1-indexed for user display
    result.metadata["columns_detected"] = [h for h in headers if h]
    
    # Detect column mapping
    col_mapping = detect_column_mapping(headers)
    
    # Check for missing critical columns
    critical_cols = ["task_name", "start_date"]
    for col in critical_cols:
        if col not in col_mapping:
            result.metadata["missing_columns"].append(col)
            result.add_warning(header_row_idx + 1, f"Missing expected column: {col}")
    
    # Track for duplicate detection
    seen_ids = set()
    valid_task_ids = set()
    source_id_map = {}
    source_sequence = 0
    current_phase = None

    # First pass: collect all task IDs for predecessor validation.
    # Project Plan 365 can export predecessor numbers without exporting the ID
    # column, so fall back to parsed source sequence -> generated prog ID.
    task_id_idx = col_mapping.get("task_id")
    for row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
        if row_idx <= header_row_idx or is_blank_row(row):
            continue
        if is_section_header_row(row, col_mapping):
            continue

        source_sequence += 1

        explicit_task_id = None
        if task_id_idx is not None and task_id_idx < len(row):
            explicit_task_id = str(row[task_id_idx]).strip() if row[task_id_idx] else None

        generated_task_id = explicit_task_id or f"prog-{source_sequence:03d}"
        valid_task_ids.add(generated_task_id)

        if explicit_task_id:
            valid_task_ids.add(explicit_task_id)
            source_id_map[explicit_task_id] = generated_task_id

        source_id_map[str(source_sequence)] = generated_task_id

    # Second pass: parse rows
    source_sequence = 0
    for row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
        result.metadata["total_rows_read"] += 1
        
        # Skip blank rows
        if is_blank_row(row):
            result.metadata["blank_rows_skipped"] += 1
            continue
        
        # Check for section headers
        if is_section_header_row(row, col_mapping):
            # Extract phase name from this row
            task_name_idx = col_mapping.get("task_name", col_mapping.get("task_id", 0))
            if task_name_idx < len(row) and row[task_name_idx]:
                current_phase = str(row[task_name_idx]).strip()
            continue
        
        # Extract values
        def get_val(field: str) -> Any:
            idx = col_mapping.get(field)
            if idx is not None and idx < len(row):
                return row[idx]
            return None
        
        source_sequence += 1

        # Task ID
        task_id = str(get_val("task_id") or "").strip()
        if not task_id:
            # Generate an ID if missing. Keep it aligned to source sequence
            # so Project Plan predecessor references can be mapped reliably.
            task_id = f"prog-{source_sequence:03d}"

        # Check for duplicates
        if task_id in seen_ids:
            result.metadata["duplicate_rows_found"] += 1
            result.add_warning(row_idx, f"Duplicate task ID: {task_id}")
            continue
        seen_ids.add(task_id)
        
        # Task name
        task_name = str(get_val("task_name") or "").strip()
        if not task_name:
            result.add_warning(row_idx, "Missing task name")
            task_name = f"Task {task_id}"
        
        # Dates
        start_date = parse_date(get_val("start_date"))
        end_date = parse_date(get_val("end_date"))
        
        if start_date:
            result.metadata["date_formats_found"].add("detected")
        
        # Duration
        duration = parse_duration(get_val("duration"), start_date, end_date)
        if duration is None and start_date and not end_date:
            duration = 1  # Default to 1 day
            result.add_warning(row_idx, f"Missing duration and end date for {task_id}, defaulting to 1 day")
        elif duration is None:
            duration = 1
            result.add_warning(row_idx, f"Could not determine duration for {task_id}, defaulting to 1 day")
        
        # Calculate end date if missing
        if start_date and not end_date and duration:
            try:
                start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                end_dt = start_dt + timedelta(days=int(duration) - 1)
                end_date = end_dt.strftime("%Y-%m-%d")
            except Exception:
                pass
        
        # Predecessors
        predecessor_raw = get_val("predecessor")
        valid_preds, invalid_preds = parse_predecessor(predecessor_raw, valid_task_ids, source_id_map)

        # FITOUTOS / PROGRAMME SELF DEPENDENCY DROP V7
        # Project Plan exports can contain predecessor references that resolve
        # back to the same generated programme ID. Drop those links rather than
        # creating circular/self dependencies in FitoutOS.
        if task_id in valid_preds:
            valid_preds = [pred for pred in valid_preds if pred != task_id]
            result.add_warning(row_idx, f"Dropped self predecessor for {task_id}")

        if invalid_preds:
            result.add_warning(row_idx, f"Invalid predecessors for {task_id}: {', '.join(invalid_preds)}")

        # Phase (from column or current section)
        phase = str(get_val("phase") or current_phase or "").strip() or "General"
        
        # Trade
        trade = str(get_val("trade") or "").strip() or "General"
        
        # Notes
        notes = str(get_val("notes") or "").strip()

        row_type, generates_task = classify_programme_row(
            task_name,
            predecessor_raw,
            get_val("trade"),
            get_val("duration"),
            get_val("hours_quoted"),
            get_val("crew_label"),
            source_sequence,
        )

        # Create the programme item
        item = {
            "id": task_id,
            "name": task_name,
            "phase": phase,
            "trade": trade,
            "duration": duration,
            "duration_unit": "days",
            "planned_start": start_date,
            "planned_finish": end_date,
            "depends_on": valid_preds,
            "task_code_id": None,
            "crew_size": None,
            "hours_per_day": 8.0,
            "resource_name": None,
            "notes": notes,
            "confidence": "parsed",
            "source": "file",
            "source_row": row_idx,
            "source_sequence": source_sequence,
            "row_type": row_type,
            "generates_task": generates_task,
            "is_summary": row_type in {"project_summary", "section", "summary"},
            "is_milestone": row_type == "milestone",
            "source_predecessors": split_predecessor_tokens(predecessor_raw),
            "original_predecessor": str(predecessor_raw).strip() if predecessor_raw is not None else "",
            "predecessor_parse_source": "source_sequence_fallback" if source_id_map else "task_id",
            "hours_quoted": _programme_number_value(get_val("hours_quoted")),
            "crew_label": get_val("crew_label"),
            "invalid_predecessors": invalid_preds,  # Track for UI warning
        }
        
        result.items.append(item)
    
    # Validation warnings
    if not result.items:
        result.add_error(0, "No valid programme items found in file")
    elif len(result.items) < 3:
        result.add_warning(0, f"Only {len(result.items)} items found - check if file was parsed correctly")
    
    # Check for missing dates
    items_without_dates = [i for i in result.items if not i["planned_start"]]
    if items_without_dates:
        result.add_warning(0, f"{len(items_without_dates)} items have no start date")
    
    return result


def parse_uploaded_file(filepath: str, filename: str) -> Dict[str, Any]:
    """
    Parse an uploaded programme file and return structured data.
    This is the main function to be called from the API.
    """
    result = parse_programme_file(filepath)
    
    return {
        "filename": filename,
        "parse_result": result.to_dict(),
        "success": len(result.items) > 0 and len(result.errors) == 0,
        "item_count": len(result.items),
        "warning_count": len(result.warnings),
        "error_count": len(result.errors),
    }

